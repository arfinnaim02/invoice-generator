import os
from flask import Flask, request, render_template, send_from_directory
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from xlsx2html import xlsx2html
import pdfkit

app = Flask(__name__)

# Ensure upload and output folders exist
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
OUTPUT_FOLDER = os.path.join(os.getcwd(), 'outputs')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER

# PDFKit config for PythonAnywhere
PDFKIT_CONFIG = pdfkit.configuration(wkhtmltopdf='/usr/bin/wkhtmltopdf')

# ------------------------ Generate Excel & PDF ------------------------
def generate_invoice(csv_path):
    # === Step 1: Load CSV ===
    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    df.columns = df.columns.str.strip()

    # === Step 2: Filter & sort columns ===
    columns_needed = ['Invoice', 'Customer Name', 'Product Name', 'Product Qty']
    df_filtered = df[columns_needed].copy()
    df_sorted = df_filtered.sort_values(by='Invoice').reset_index(drop=True)
    df_sorted.insert(0, 'Serial No.', df_sorted.index + 1)

    # === Step 3: Save to Excel with styling ===
    today = datetime.now().strftime("%Y-%m-%d")
    excel_filename = f"Vibes_Invoice_{today}.xlsx"
    excel_path = os.path.join(app.config['OUTPUT_FOLDER'], excel_filename)
    df_sorted.to_excel(excel_path, index=False, startrow=1)

    # Styling with openpyxl
    wb = load_workbook(excel_path)
    ws = wb.active
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    header_font = Font(bold=True)

    total_columns = len(df_sorted.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws['A1'] = f"Vibes - {today}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment
    ws['A1'].fill = header_fill

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=total_columns):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = border
        ws.row_dimensions[row[0].row].height = 25

    for row in range(1, ws.max_row + 1):
        for col in range(1, total_columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border

    for col_idx in range(1, total_columns + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            len(str(cell.value)) 
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx) 
            for cell in row if cell.value
        )
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(excel_path)

    # === Step 4: Export to PDF using xlsx2html + pdfkit ===
    pdf_filename = excel_filename.replace(".xlsx", ".pdf")
    pdf_path = os.path.join(app.config['OUTPUT_FOLDER'], pdf_filename)

    # Convert Excel → HTML → PDF
    html_path = excel_path.replace(".xlsx", ".html")
    with open(html_path, "w", encoding="utf-8") as f:
        xlsx2html(excel_path, f)

    # Convert HTML to PDF
    pdfkit.from_file(html_path, pdf_path, configuration=PDFKIT_CONFIG)

    return excel_filename, pdf_filename, df_sorted

# ------------------------ Routes ------------------------
@app.route("/", methods=["GET", "POST"])
def index():
    excel_file = pdf_file = None
    preview_data = None
    error = None

    if request.method == "POST":
        if 'file' not in request.files or request.files['file'].filename == '':
            error = "No file selected"
        else:
            uploaded_file = request.files['file']
            upload_path = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file.filename)
            uploaded_file.save(upload_path)

            try:
                excel_file, pdf_file, preview_data = generate_invoice(upload_path)
            except Exception as e:
                error = f"Error processing file: {e}"

    return render_template(
        "index.html",
        excel_file=excel_file,
        pdf_file=pdf_file,
        preview_data=preview_data,
        error=error
    )

@app.route("/downloads/<filename>")
def download_file(filename):
    return send_from_directory(app.config['OUTPUT_FOLDER'], filename, as_attachment=True)

if __name__ == "__main__":
    # For local testing only; set debug=False when deploying
    app.run(debug=False, host='0.0.0.0', port=5000)
